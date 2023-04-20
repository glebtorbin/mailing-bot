from openpyxl import load_workbook
from .wa_api import LOGGER


async def WA_xlsx_search(keyword, user_id):
    try:
        file = f"./wa_mailing_contacts/wa_{user_id}.xlsx"
        wb = load_workbook(file, read_only=True)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=3, max_col=30):
            for cell in row:
                if cell.value == keyword:
                    print(cell.value)
                    # if (sheet.cell(row=cell.row + 1, column=cell.column).value).isdigit():
                    phone_row = sheet.cell(row=cell.row, column=cell.column).row
                    phone_col = sheet.cell(row=cell.row, column=cell.column).column
                    break
        print(sheet.max_row)
        num_list = [sheet.cell(row=s, column=phone_col).value for s in range(2, sheet.max_row + 1)]
        return num_list, len(num_list)
    except Exception as err:
        LOGGER.error(err)
